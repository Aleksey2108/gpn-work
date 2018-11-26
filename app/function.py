# -*- coding: utf-8 -*-
import os, datetime, webbrowser
from app import app, db
from flask import  url_for, redirect
from app.models import AuditTrail, AuditTrail_CHS, AuditTrail_GO, AuditTrail_PB
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side, NamedStyle
from flask import send_from_directory



fill = PatternFill(fill_type='solid',
                   start_color='c1c1c1',
                   end_color='c2c2c2')
fill_1 = PatternFill(fill_type='solid',
                   start_color='ebf1de',
                   end_color='ebf1de')

fill_2 = PatternFill(fill_type='solid',
                   start_color='d9f1ff',
                   end_color='d9f1ff')

fill_3 = PatternFill(fill_type='solid',
                   start_color='ffff00',
                   end_color='ffff00')

fill_4 = PatternFill(fill_type='solid',
                   start_color='b8cce4',
                   end_color='b8cce4')


fill_5 = PatternFill(fill_type='solid',
                   start_color='f2f2f2',
                   end_color='f2f2f2')


align_center=Alignment(horizontal='center',
                       vertical='bottom',
                       text_rotation=0,
                       wrap_text=True,
                       shrink_to_fit=False,
                       indent=0)
align_center1=Alignment(horizontal='center',
                       vertical='center',
                       text_rotation=0,
                       wrap_text=True,
                       shrink_to_fit=False,
                       indent=0)
align_left=Alignment(horizontal='left',
                       vertical='bottom',
                       text_rotation=0,
#                       wrap_text=False,
                       wrap_text=True,
                       shrink_to_fit=False,
                       indent=0)
align_right=Alignment(horizontal='right',
                       vertical='bottom',
                       text_rotation=0,
                       wrap_text=False,
                       shrink_to_fit=False,
                       indent=0)
text_vertical =Alignment(horizontal='center',
                       vertical='center',
                       text_rotation=90,
                       wrap_text=True,
                       shrink_to_fit=False,
                       indent=0)
font1 = Font(name='Times New Roman',
                    size=12,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
font2 = Font(name='Times New Roman',
                    size=14,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
font3 = Font(name='Times New Roman',
                    size=16,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
font4 = Font(name='Times New Roman',
                    size=14,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
font5 = Font(name='Times New Roman',
                    size=7,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
font6 = Font(name='Times New Roman',
                    size=12,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FFFF0000')
font7 = Font(name='Times New Roman',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
border = Border(left=Side(border_style='thin',
                             color='FF000000'),
                   right=Side(border_style='thin',
                              color='FF000000'),
                   top=Side(border_style='thin',
                            color='FF000000'),
                   bottom=Side(border_style='thin',
                               color='FF000000'),
                   diagonal=Side(border_style='thin',
                                 color='FF000000'),
                   diagonal_direction=0,
                   outline=Side(border_style='thin',
                                color='FF000000'),
                   vertical=Side(border_style='thin',
                                 color='FF000000'),
                   horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )

border1 = Border(left=Side(border_style='thin',
                             color='FF000000'),
                   top=Side(border_style='thin',
                            color='FF000000'),
                   bottom=Side(border_style='thin',
                               color='FF000000'),
                   diagonal=Side(border_style='thin',
                                 color='FF000000'),
                   diagonal_direction=0,
                   outline=Side(border_style='thin',
                                color='FF000000'),
                   vertical=Side(border_style='thin',
                                 color='FF000000'),
                   horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )

border2 = Border(right=Side(border_style='thin',
                              color='FF000000'),
                   top=Side(border_style='thin',
                            color='FF000000'),
                   bottom=Side(border_style='thin',
                               color='FF000000'),
                   diagonal=Side(border_style='thin',
                                 color='FF000000'),
                   diagonal_direction=0,
                   outline=Side(border_style='thin',
                                color='FF000000'),
                   vertical=Side(border_style='thin',
                                 color='FF000000'),
                   horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )
protection = Protection(locked=True,
                          hidden=False)

def GetDepartment(depart_id):
    depart = [
         {
              'name' : u'УНПР',
              'id' : 'unpr'
         },
         {
              'name' : u'НТО',
              'id' : 'nto'
         },
         {
              'name' : u'ОНБП',
              'id' : 'onbp'
         },
         {
              'name' : u'ОНГОЗНТЧС',
              'id' : 'ongozntchs'
         },
         {
              'name' : u'ОГСД',
              'id' : 'ogsd'
         },
         {
              'name' : u'ООРД',
              'id' : 'oord'
         },
         {
              'name' : u'ОЛК',
              'id' : 'olk'
         },
         {
              'name' : u'ОНОВПО',
              'id' : 'onovpo'
         },
         {
              'name' : u'ОНТ',
              'id' : 'ont'
         },
         {
              'name' : u'ОПД',
              'id' : 'opd'
         },
         {
              'name' : u'РОНПР',
              'id' : 'ronpr'
         },
    ]
    if depart_id == 'null':
       return depart
    else:
       for val in depart:
         if val['id'] == depart_id:
           return val['name']

def GetDepartmentAraay():

    aray_t = GetDepartment('null')
    for val in aray_t:
      array = "('" + val['id'] +"','" + val['name'] + "'),"
    return array

def GetLastDay(month, year):
    if month == 2:
      if year == 2020 or year == 2024 or year == 2028:
         day = 29
      else:
         day = 28
    elif month == 1 or month == 3 or month == 5 or month == 7 or month == 8 or month == 10 or month == 12:
       day = 31
    else: 
       day = 30
    return day

def CheckLastDay(day, month, year):
    if day <= 28:
       return day

    if month == 2:
      if year == 2020 or year == 2024 or year == 2028:
         day = 29
    elif month == 1 or month == 3 or month == 5 or month == 7 or month == 8 or month == 10 or month == 12:
       if  day > 31:
         day = 31
    else: 
       if day > 30:
         day = 30
    return day

def CreateAuditTrailXls_CHS(start_date, end_date):

# !!  В запрос необходимо добавить условие выбора по отделу !!
    rows =  AuditTrail_CHS.query.filter(start_date <= AuditTrail.checkdate,  AuditTrail.checkdate <= end_date).all()
    
    if rows:
#      wb = Workbook()
      load_filename = '%s/%s'  % (app.config['MASTER_FOLDER'], 'AT_ZNTCHS.xlsx')
      wb = load_workbook(load_filename)
      ws1 = wb.worksheets[0] 

      ws1['A17'] =  u' Начат: "01" января ' + str(start_date.year) + u' года'
      ws1['A19'] =  u' Окончен: "      " ____________ ' + str(end_date.year) + u' года'
 
      ws2 = wb.worksheets[1] 

      key = 4
      n_pp =1

      for row in rows:
         ws2.cell(row=key, column=1).value = n_pp
         ws2.cell(row=key, column=1).font = font4
         ws2.cell(row=key, column=1).alignment = align_center1
         ws2.cell(row=key, column=1).border = border
         ws2.cell(row=key, column=2).value = row.objectname
         ws2.cell(row=key, column=2).font = font4
         ws2.cell(row=key, column=2).alignment = align_center1
         ws2.cell(row=key, column=2).border = border
         ws2.cell(row=key, column=3).value = u'г. Москва, ' + row.objectadres
         ws2.cell(row=key, column=3).font = font4
         ws2.cell(row=key, column=3).alignment = align_center1
         ws2.cell(row=key, column=3).border = border
         ws2.cell(row=key, column=4).value = row.doc_stored
         ws2.cell(row=key, column=4).font = font4
         ws2.cell(row=key, column=4).alignment = align_center1
         ws2.cell(row=key, column=4).border = border

         ws2.cell(row=key, column=5).value = n_pp
         ws2.cell(row=key, column=5).font = font4
         ws2.cell(row=key, column=5).alignment = align_center1
         ws2.cell(row=key, column=5).border = border1
         ws2.cell(row=key, column=5).fill = fill_5

         ws2.cell(row=key, column=6).value = row.checkdate
         ws2.cell(row=key, column=6).font = font4
         ws2.cell(row=key, column=6).alignment = align_center1
         ws2.cell(row=key, column=6).border = border2


         ws2.cell(row=key, column=7).value = '%s \n %s \n %s \n'  % (row.type_inspection, str(row.start_date), str(row.end_date)) 
         ws2.cell(row=key, column=7).font = font4
         ws2.cell(row=key, column=7).alignment = align_center1
         ws2.cell(row=key, column=7).border = border

         if row.act_number:
            ws2.cell(row=key, column=8).value = '%s \n %s \n'  % (row.act_number, str(row.act_date)) 
         ws2.cell(row=key, column=8).font = font4
         ws2.cell(row=key, column=8).alignment = align_center1
         ws2.cell(row=key, column=8).border = border

         if row.order_number:
            ws2.cell(row=key, column=9).value = '%s \n %s \n'  % (row.order_number, str(row.order_date)) 
         ws2.cell(row=key, column=9).font = font4
         ws2.cell(row=key, column=9).alignment = align_center1
         ws2.cell(row=key, column=9).border = border

         ws2.cell(row=key, column=10).value = row.of_violations
         ws2.cell(row=key, column=10).font = font4
         ws2.cell(row=key, column=10).alignment = align_center1
         ws2.cell(row=key, column=10).border = border
         ws2.cell(row=key, column=11).value = row.of_violations_unscheduled
         ws2.cell(row=key, column=11).font = font4
         ws2.cell(row=key, column=11).alignment = align_center1
         ws2.cell(row=key, column=11).border = border
         ws2.cell(row=key, column=12).value = row.fixed_violations
         ws2.cell(row=key, column=12).font = font4
         ws2.cell(row=key, column=12).alignment = align_center1
         ws2.cell(row=key, column=12).border = border

         ws2.cell(row=key, column=13).value = row.name_employee
         ws2.cell(row=key, column=13).font = font4
         ws2.cell(row=key, column=13).alignment = align_center1
         ws2.cell(row=key, column=13).fill = fill_1
         ws2.cell(row=key, column=13).border = border


         ws2.cell(row=key, column=14).value = row.depart_name
         ws2.cell(row=key, column=14).font = font4
         ws2.cell(row=key, column=14).alignment = align_center1
         ws2.cell(row=key, column=14).fill = fill
         ws2.cell(row=key, column=14).border = border

         if row.check_number:
            ws2.cell(row=key, column=15).value = row.check_number
         ws2.cell(row=key, column=15).font = font4
         ws2.cell(row=key, column=15).alignment = align_center1
         ws2.cell(row=key, column=15).fill = fill_2
         ws2.cell(row=key, column=15).border = border



         key = key + 1
         n_pp = n_pp + 1
     
      range = '%s2:%s%s' %('A' , 'N' , key-1)
      print 'range = ' + str(range)

      ws2.auto_filter.ref = str(range)
#      ws2.auto_filter.add_sort_condition("A4:A8")

      cell_num = key + 1
      ws2.cell(row=cell_num, column=1).value = u'Всего'
      ws2.cell(row=cell_num, column=1).font = font5
      ws2.cell(row=cell_num, column=1).alignment = align_center1
      ws2.cell(row=cell_num, column=1).fill = fill_3
      ws2.cell(row=cell_num, column=1).border = border

      ws2.cell(row=cell_num, column=2).fill = fill_3
      ws2.cell(row=cell_num, column=2).border = border
      ws2.cell(row=cell_num, column=3).fill = fill_3
      ws2.cell(row=cell_num, column=3).border = border
      ws2.cell(row=cell_num, column=4).fill = fill_3
      ws2.cell(row=cell_num, column=4).border = border
      ws2.cell(row=cell_num, column=5).fill = fill_3
      ws2.cell(row=cell_num, column=5).border = border
      ws2.cell(row=cell_num, column=6).fill = fill_3
      ws2.cell(row=cell_num, column=6).border = border

      formula = '=SUM(J4:J' + str(key-1) + ')'
      ws2.cell(row=cell_num, column=10).value = formula
      ws2.cell(row=cell_num, column=10).font = font4
      ws2.cell(row=cell_num, column=10).alignment = align_center1
      ws2.cell(row=cell_num, column=10).fill = fill_3
      ws2.cell(row=cell_num, column=10).border = border

      formula = '=SUM(K4:K' + str(key-1) + ')'
      ws2.cell(row=cell_num, column=11).value = formula
      ws2.cell(row=cell_num, column=11).font = font4
      ws2.cell(row=cell_num, column=11).alignment = align_center1
      ws2.cell(row=cell_num, column=11).fill = fill_3
      ws2.cell(row=cell_num, column=11).border = border

      formula = '=SUM(L4:L' + str(key-1) + ')'
      ws2.cell(row=cell_num, column=12).value = formula
      ws2.cell(row=cell_num, column=12).font = font4
      ws2.cell(row=cell_num, column=12).alignment = align_center1
      ws2.cell(row=cell_num, column=12).fill = fill_3
      ws2.cell(row=cell_num, column=12).border = border

      ws2.cell(row=cell_num, column=7).fill = fill_3
      ws2.cell(row=cell_num, column=7).border = border
      ws2.cell(row=cell_num, column=8).fill = fill_3
      ws2.cell(row=cell_num, column=8).border = border
      ws2.cell(row=cell_num, column=9).fill = fill_3
      ws2.cell(row=cell_num, column=9).border = border
      ws2.cell(row=cell_num, column=13).fill = fill_3
      ws2.cell(row=cell_num, column=13).border = border
      ws2.cell(row=cell_num, column=14).fill = fill_3
      ws2.cell(row=cell_num, column=14).border = border
      ws2.cell(row=cell_num, column=15).fill = fill_3
      ws2.cell(row=cell_num, column=15).border = border

      now = datetime.datetime.now()

#      print 'Folder =' + app.config['UPLOAD_FOLDER']
      filename = 'svao1_chs_' +str(now.year)+'-'+str(now.month)+'-'+str(now.day)+'_'+str(now.hour)+'-'+str(now.minute)+'-'+str(now.second)+'.xlsx'
#      dest_filename = 'svao1_chs_' +str(now.year)+'-'+str(now.month)+'-'+str(now.day)+'_'+str(now.hour)+'-'+str(now.minute)+'-'+str(now.second)+'.xlsx'
      dest_filename ='%s/%s'  % (app.config['UPLOAD_FOLDER'], filename)

      wb.save(filename = dest_filename)

      return filename

    else:
      result = 'error'
      return result

def CreateAuditTrailXls_GO(start_date, end_date):

# !!  В запрос необходимо добавить условие выбора по отделу !! 
    rows =  AuditTrail_GO.query.filter(start_date <= AuditTrail.checkdate,  AuditTrail.checkdate <= end_date).all()
    
    if rows:
      load_filename = '%s/%s'  % (app.config['MASTER_FOLDER'], 'AT_GO.xlsx')
      wb = load_workbook(load_filename)
      ws1 = wb.worksheets[0] 


      ws1['A17'] =  u' Начат: "01" января ' + str(start_date.year) + u' года'
      ws1['A19'] =  u' Окончен: "      " ____________ ' + str(end_date.year) + u' года'
 
      ws2 = wb.worksheets[1] 


      key = 4
      n_pp =1

      for row in rows:
         ws2.cell(row=key, column=1).value = n_pp
         ws2.cell(row=key, column=1).font = font4
         ws2.cell(row=key, column=1).alignment = align_center1
         ws2.cell(row=key, column=1).border = border
         ws2.cell(row=key, column=2).value = row.objectname
         ws2.cell(row=key, column=2).font = font4
         ws2.cell(row=key, column=2).alignment = align_center1
         ws2.cell(row=key, column=2).border = border
         ws2.cell(row=key, column=3).value = u'г. Москва, ' + row.objectadres
         ws2.cell(row=key, column=3).font = font4
         ws2.cell(row=key, column=3).alignment = align_center1
         ws2.cell(row=key, column=3).border = border
         ws2.cell(row=key, column=4).value = row.doc_stored
         ws2.cell(row=key, column=4).font = font4
         ws2.cell(row=key, column=4).alignment = align_center1
         ws2.cell(row=key, column=4).border = border

         ws2.cell(row=key, column=5).value = n_pp
         ws2.cell(row=key, column=5).font = font4
         ws2.cell(row=key, column=5).alignment = align_center1
         ws2.cell(row=key, column=5).border = border1
         ws2.cell(row=key, column=5).fill = fill_5

         ws2.cell(row=key, column=6).value = row.checkdate
         ws2.cell(row=key, column=6).font = font4
         ws2.cell(row=key, column=6).alignment = align_center1
         ws2.cell(row=key, column=6).border = border2


         ws2.cell(row=key, column=7).value = '%s \n %s \n %s \n'  % (row.type_inspection, str(row.start_date), str(row.end_date)) 
         ws2.cell(row=key, column=7).font = font4
         ws2.cell(row=key, column=7).alignment = align_center1
         ws2.cell(row=key, column=7).border = border

         if row.act_number:
            ws2.cell(row=key, column=8).value = '%s \n %s \n'  % (row.act_number, str(row.act_date)) 
         ws2.cell(row=key, column=8).font = font4
         ws2.cell(row=key, column=8).alignment = align_center1
         ws2.cell(row=key, column=8).border = border

         if row.order_number:
            ws2.cell(row=key, column=9).value = '%s \n %s \n'  % (row.order_number, str(row.order_date)) 
         ws2.cell(row=key, column=9).font = font4
         ws2.cell(row=key, column=9).alignment = align_center1
         ws2.cell(row=key, column=9).border = border

         ws2.cell(row=key, column=10).value = row.of_violations
         ws2.cell(row=key, column=10).font = font4
         ws2.cell(row=key, column=10).alignment = align_center1
         ws2.cell(row=key, column=10).border = border
         ws2.cell(row=key, column=11).value = row.of_violations_unscheduled
         ws2.cell(row=key, column=11).font = font4
         ws2.cell(row=key, column=11).alignment = align_center1
         ws2.cell(row=key, column=11).border = border
         ws2.cell(row=key, column=12).value = row.fixed_violations
         ws2.cell(row=key, column=12).font = font4
         ws2.cell(row=key, column=12).alignment = align_center1
         ws2.cell(row=key, column=12).border = border

         ws2.cell(row=key, column=13).value = row.name_employee
         ws2.cell(row=key, column=13).font = font4
         ws2.cell(row=key, column=13).alignment = align_center1
         ws2.cell(row=key, column=13).fill = fill_1
         ws2.cell(row=key, column=13).border = border

         ws2.cell(row=key, column=14).value = row.other_documents
         ws2.cell(row=key, column=14).font = font4
         ws2.cell(row=key, column=14).alignment = align_center1
         ws2.cell(row=key, column=14).border = border


         ws2.cell(row=key, column=15).value = row.depart_name
         ws2.cell(row=key, column=15).font = font4
         ws2.cell(row=key, column=15).alignment = align_center1
         ws2.cell(row=key, column=15).fill = fill
         ws2.cell(row=key, column=15).border = border

         if row.check_number:
            ws2.cell(row=key, column=16).value = row.check_number
         ws2.cell(row=key, column=16).font = font4
         ws2.cell(row=key, column=16).alignment = align_center1
         ws2.cell(row=key, column=16).fill = fill_2
         ws2.cell(row=key, column=16).border = border


         key = key + 1
         n_pp = n_pp + 1
     
      range = '%s2:%s%s' %('A' , 'N' , key-1)
      print 'range = ' + str(range)

      ws2.auto_filter.ref = str(range)
#      ws2.auto_filter.add_sort_condition("A4:A8")

      cell_num = key + 1
      ws2.cell(row=cell_num, column=1).value = u'Всего'
      ws2.cell(row=cell_num, column=1).font = font5
      ws2.cell(row=cell_num, column=1).alignment = align_center1
      ws2.cell(row=cell_num, column=1).fill = fill_3
      ws2.cell(row=cell_num, column=1).border = border

      ws2.cell(row=cell_num, column=2).fill = fill_3
      ws2.cell(row=cell_num, column=2).border = border
      ws2.cell(row=cell_num, column=3).fill = fill_3
      ws2.cell(row=cell_num, column=3).border = border
      ws2.cell(row=cell_num, column=4).fill = fill_3
      ws2.cell(row=cell_num, column=4).border = border
      ws2.cell(row=cell_num, column=5).fill = fill_3
      ws2.cell(row=cell_num, column=5).border = border
      ws2.cell(row=cell_num, column=6).fill = fill_3
      ws2.cell(row=cell_num, column=6).border = border

      formula = '=SUM(J4:J' + str(key-1) + ')'
      ws2.cell(row=cell_num, column=10).value = formula
      ws2.cell(row=cell_num, column=10).font = font4
      ws2.cell(row=cell_num, column=10).alignment = align_center1
      ws2.cell(row=cell_num, column=10).fill = fill_3
      ws2.cell(row=cell_num, column=10).border = border

      formula = '=SUM(K4:K' + str(key-1) + ')'
      ws2.cell(row=cell_num, column=11).value = formula
      ws2.cell(row=cell_num, column=11).font = font4
      ws2.cell(row=cell_num, column=11).alignment = align_center1
      ws2.cell(row=cell_num, column=11).fill = fill_3
      ws2.cell(row=cell_num, column=11).border = border

      formula = '=SUM(L4:L' + str(key-1) + ')'
      ws2.cell(row=cell_num, column=12).value = formula
      ws2.cell(row=cell_num, column=12).font = font4
      ws2.cell(row=cell_num, column=12).alignment = align_center1
      ws2.cell(row=cell_num, column=12).fill = fill_3
      ws2.cell(row=cell_num, column=12).border = border

      ws2.cell(row=cell_num, column=7).fill = fill_3
      ws2.cell(row=cell_num, column=7).border = border
      ws2.cell(row=cell_num, column=8).fill = fill_3
      ws2.cell(row=cell_num, column=8).border = border
      ws2.cell(row=cell_num, column=9).fill = fill_3
      ws2.cell(row=cell_num, column=9).border = border
      ws2.cell(row=cell_num, column=13).fill = fill_3
      ws2.cell(row=cell_num, column=13).border = border
      ws2.cell(row=cell_num, column=14).fill = fill_3
      ws2.cell(row=cell_num, column=14).border = border
      ws2.cell(row=cell_num, column=15).fill = fill_3
      ws2.cell(row=cell_num, column=15).border = border
      ws2.cell(row=cell_num, column=16).fill = fill_3
      ws2.cell(row=cell_num, column=16).border = border

      now = datetime.datetime.now()

      filename = 'svao1_go_' +str(now.year)+'-'+str(now.month)+'-'+str(now.day)+'_'+str(now.hour)+'-'+str(now.minute)+'-'+str(now.second)+'.xlsx'
      dest_filename ='%s/%s'  % (app.config['UPLOAD_FOLDER'], filename)

      wb.save(filename = dest_filename)
      return filename

    else:
      result = 'error'
      return result

def CreateAuditTrailXls_PB(start_date, end_date):

# !!  В запрос необходимо добавить условие выбора по отделу !! 
    rows =  AuditTrail_PB.query.filter(start_date <= AuditTrail.checkdate,  AuditTrail.checkdate <= end_date).all()
    
    if rows:
      load_filename = '%s/%s'  % (app.config['MASTER_FOLDER'], 'AT_GO.xlsx')
      wb = load_workbook(load_filename)
      ws1 = wb.worksheets[0] 


      ws1['A17'] =  u' Начат: "01" января ' + str(start_date.year) + u' года'
      ws1['A19'] =  u' Окончен: "      " ____________ ' + str(end_date.year) + u' года'
 
      ws2 = wb.worksheets[1] 


      key = 4
      n_pp =1

      for row in rows:
         ws2.cell(row=key, column=1).value = n_pp
         ws2.cell(row=key, column=1).font = font4
         ws2.cell(row=key, column=1).alignment = align_center1
         ws2.cell(row=key, column=1).border = border
         ws2.cell(row=key, column=2).value = row.objectname
         ws2.cell(row=key, column=2).font = font4
         ws2.cell(row=key, column=2).alignment = align_center1
         ws2.cell(row=key, column=2).border = border
         ws2.cell(row=key, column=3).value = u'г. Москва, ' + row.objectadres
         ws2.cell(row=key, column=3).font = font4
         ws2.cell(row=key, column=3).alignment = align_center1
         ws2.cell(row=key, column=3).border = border
         ws2.cell(row=key, column=4).value = row.doc_stored
         ws2.cell(row=key, column=4).font = font4
         ws2.cell(row=key, column=4).alignment = align_center1
         ws2.cell(row=key, column=4).border = border

         ws2.cell(row=key, column=5).value = n_pp
         ws2.cell(row=key, column=5).font = font4
         ws2.cell(row=key, column=5).alignment = align_center1
         ws2.cell(row=key, column=5).border = border1
         ws2.cell(row=key, column=5).fill = fill_5

         ws2.cell(row=key, column=6).value = row.checkdate
         ws2.cell(row=key, column=6).font = font4
         ws2.cell(row=key, column=6).alignment = align_center1
         ws2.cell(row=key, column=6).border = border2


         ws2.cell(row=key, column=7).value = '%s \n %s \n %s \n'  % (row.type_inspection, str(row.start_date), str(row.end_date)) 
         ws2.cell(row=key, column=7).font = font4
         ws2.cell(row=key, column=7).alignment = align_center1
         ws2.cell(row=key, column=7).border = border

         if row.act_number:
            ws2.cell(row=key, column=8).value = '%s \n %s \n'  % (row.act_number, str(row.act_date)) 
         ws2.cell(row=key, column=8).font = font4
         ws2.cell(row=key, column=8).alignment = align_center1
         ws2.cell(row=key, column=8).border = border

         if row.order_number:
            ws2.cell(row=key, column=9).value = '%s \n %s \n'  % (row.order_number, str(row.order_date)) 
         ws2.cell(row=key, column=9).font = font4
         ws2.cell(row=key, column=9).alignment = align_center1
         ws2.cell(row=key, column=9).border = border

         ws2.cell(row=key, column=10).value = row.of_violations
         ws2.cell(row=key, column=10).font = font4
         ws2.cell(row=key, column=10).alignment = align_center1
         ws2.cell(row=key, column=10).border = border
         ws2.cell(row=key, column=11).value = row.of_violations_unscheduled
         ws2.cell(row=key, column=11).font = font4
         ws2.cell(row=key, column=11).alignment = align_center1
         ws2.cell(row=key, column=11).border = border
         ws2.cell(row=key, column=12).value = row.fixed_violations
         ws2.cell(row=key, column=12).font = font4
         ws2.cell(row=key, column=12).alignment = align_center1
         ws2.cell(row=key, column=12).border = border

         ws2.cell(row=key, column=13).value = row.name_employee
         ws2.cell(row=key, column=13).font = font4
         ws2.cell(row=key, column=13).alignment = align_center1
         ws2.cell(row=key, column=13).fill = fill_1
         ws2.cell(row=key, column=13).border = border

         ws2.cell(row=key, column=14).value = row.other_documents
         ws2.cell(row=key, column=14).font = font4
         ws2.cell(row=key, column=14).alignment = align_center1
         ws2.cell(row=key, column=14).border = border


         ws2.cell(row=key, column=15).value = row.depart_name
         ws2.cell(row=key, column=15).font = font4
         ws2.cell(row=key, column=15).alignment = align_center1
         ws2.cell(row=key, column=15).fill = fill
         ws2.cell(row=key, column=15).border = border

         if row.check_number:
            ws2.cell(row=key, column=16).value = row.check_number
         ws2.cell(row=key, column=16).font = font4
         ws2.cell(row=key, column=16).alignment = align_center1
         ws2.cell(row=key, column=16).fill = fill_2
         ws2.cell(row=key, column=16).border = border


         key = key + 1
         n_pp = n_pp + 1
     
      range = '%s2:%s%s' %('A' , 'N' , key-1)
      print 'range = ' + str(range)

      ws2.auto_filter.ref = str(range)
#      ws2.auto_filter.add_sort_condition("A4:A8")

      cell_num = key + 1
      ws2.cell(row=cell_num, column=1).value = u'Всего'
      ws2.cell(row=cell_num, column=1).font = font5
      ws2.cell(row=cell_num, column=1).alignment = align_center1
      ws2.cell(row=cell_num, column=1).fill = fill_3
      ws2.cell(row=cell_num, column=1).border = border

      ws2.cell(row=cell_num, column=2).fill = fill_3
      ws2.cell(row=cell_num, column=2).border = border
      ws2.cell(row=cell_num, column=3).fill = fill_3
      ws2.cell(row=cell_num, column=3).border = border
      ws2.cell(row=cell_num, column=4).fill = fill_3
      ws2.cell(row=cell_num, column=4).border = border
      ws2.cell(row=cell_num, column=5).fill = fill_3
      ws2.cell(row=cell_num, column=5).border = border
      ws2.cell(row=cell_num, column=6).fill = fill_3
      ws2.cell(row=cell_num, column=6).border = border

      formula = '=SUM(J4:J' + str(key-1) + ')'
      ws2.cell(row=cell_num, column=10).value = formula
      ws2.cell(row=cell_num, column=10).font = font4
      ws2.cell(row=cell_num, column=10).alignment = align_center1
      ws2.cell(row=cell_num, column=10).fill = fill_3
      ws2.cell(row=cell_num, column=10).border = border

      formula = '=SUM(K4:K' + str(key-1) + ')'
      ws2.cell(row=cell_num, column=11).value = formula
      ws2.cell(row=cell_num, column=11).font = font4
      ws2.cell(row=cell_num, column=11).alignment = align_center1
      ws2.cell(row=cell_num, column=11).fill = fill_3
      ws2.cell(row=cell_num, column=11).border = border

      formula = '=SUM(L4:L' + str(key-1) + ')'
      ws2.cell(row=cell_num, column=12).value = formula
      ws2.cell(row=cell_num, column=12).font = font4
      ws2.cell(row=cell_num, column=12).alignment = align_center1
      ws2.cell(row=cell_num, column=12).fill = fill_3
      ws2.cell(row=cell_num, column=12).border = border

      ws2.cell(row=cell_num, column=7).fill = fill_3
      ws2.cell(row=cell_num, column=7).border = border
      ws2.cell(row=cell_num, column=8).fill = fill_3
      ws2.cell(row=cell_num, column=8).border = border
      ws2.cell(row=cell_num, column=9).fill = fill_3
      ws2.cell(row=cell_num, column=9).border = border
      ws2.cell(row=cell_num, column=13).fill = fill_3
      ws2.cell(row=cell_num, column=13).border = border
      ws2.cell(row=cell_num, column=14).fill = fill_3
      ws2.cell(row=cell_num, column=14).border = border
      ws2.cell(row=cell_num, column=15).fill = fill_3
      ws2.cell(row=cell_num, column=15).border = border
      ws2.cell(row=cell_num, column=16).fill = fill_3
      ws2.cell(row=cell_num, column=16).border = border


      now = datetime.datetime.now()

      filename = 'svao1_pb_' +str(now.year)+'-'+str(now.month)+'-'+str(now.day)+'_'+str(now.hour)+'-'+str(now.minute)+'-'+str(now.second)+'.xlsx'
      dest_filename ='%s/%s'  % (app.config['UPLOAD_FOLDER'], filename)

      wb.save(filename = dest_filename)
      return filename
    else:
      result = 'error'
      return result

def XlsLine(cell , value =None, alignment =None, font =None, fill = None):
  try:
     cell
  except KeyError:
     return ''
  else:
    print  'cell = '+cell
    r_date = ''
  
    if  value is not None:
      r_date = value

    if  alignment is not None:
      r_date.alignment = alignment

    if  font is not None:
      r_date.font = font

    return r_date
